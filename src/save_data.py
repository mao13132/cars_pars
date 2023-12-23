# ---------------------------------------------
# Program by @developer_telegrams
#
#
# Version   Date        Info
# 1.0       2023    Initial Version
#
# ---------------------------------------------

from openpyxl import Workbook

from settings import dir_project


class SaveData:
    def __init__(self, data_pars):
        self.columns = ['product_id', 'name', 'categories', 'sku', 'brand', 'price', 'quantity',
                        'additional information', 'analogs', 'image_name']

        self.wb = Workbook()

        self.ws = self.wb.active

        self.data_pars = data_pars

        self.title = ''

        self.subcategory = ''

        self.brand = ''

        self.count_row = 2

        self.count_product = 1

        self.last_columns = len(self.columns) + 1

        self.cords_title = {}

    def iter_article(self, data_):
        for article, data_article in data_.items():

            self.ws.cell(row=self.count_row, column=1).value = self.count_product

            self.ws.cell(row=self.count_row, column=2).value = data_article['name']

            self.ws.cell(row=self.count_row, column=3).value = f"{self.title} | {self.subcategory}"

            self.ws.cell(row=self.count_row, column=4).value = article

            self.ws.cell(row=self.count_row, column=5).value = self.brand

            self.ws.cell(row=self.count_row, column=6).value = data_article['price']

            self.ws.cell(row=self.count_row, column=7).value = data_article['quantity']

            self.ws.cell(row=self.count_row, column=8).value = data_article['more']

            self.ws.cell(row=self.count_row, column=9).value = ' | '.join(x for x in data_article['alternative'])

            self.ws.cell(row=self.count_row, column=10).value = data_article['image'].split('data\\')[-1]

            for sub_column, value in data_article['kriterii'].items():

                exist = self.cords_title.get(sub_column, False)

                target_column = self.last_columns

                if not exist:
                    self.cords_title[sub_column] = target_column

                    self.ws.cell(row=1, column=target_column).value = sub_column

                    self.last_columns += 1

                else:
                    target_column = exist

                self.ws.cell(row=self.count_row, column=target_column).value = value

            self.count_product += 1

            self.count_row += 1

            print()

    def iter_brand(self, data_):
        for name_brand, data_brand in data_.items():
            self.brand = name_brand

            res = self.iter_article(data_brand)

            print()

    def iter_subcategory(self, data_):
        for name_subcategory, data_sub in data_.items():
            self.subcategory = name_subcategory

            res_brand = self.iter_brand(data_sub)

            print()

    def iter_category(self):
        for count_category, data_category in self.data_pars.items():
            self.title = data_category['title']

            res_iter = self.iter_subcategory(data_category['subcategory'])

            print()

    def create_title(self):

        count_column = 1

        for col in range(len(self.columns)):
            self.ws.cell(row=1, column=count_column).value = self.columns[col]

            count_column += 1

        return True

    def save_data(self):

        self.create_title()

        res_ = self.iter_category()

        print()

        self.wb.save(f'{dir_project}\\test.xlsx')

        return True


if __name__ == '__main__':
    from _temp import pars_data

    data_pars = pars_data

    res = SaveData(data_pars).save_data()
